import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image





def save_results_to_excel(phase, func, eq_coef, mse, save_dir, plot):
    result_file = 'result.xlsx'
    result_file_path = os.path.join(save_dir, result_file)

    sheet_name = 'Means'
    if not os.path.exists(result_file_path):
        # Если файла нет, создать DataFrame и сохранить его в файл
        result_df = pd.DataFrame(columns=['Phase', 'Function', 'Equation', 'Mean Squared Error'])
        result_df.to_excel(result_file_path, sheet_name=sheet_name, index=False)

    result_df = pd.DataFrame(columns=['Phase', 'Function', 'Equation', 'Mean Squared Error'])

    equation = equation_generate(func, eq_coef)

    new_row = {
        'Phase': phase,
        'Function': func,
        'Equation': equation,
        'Mean Squared Error': mse
    }

    result_df = pd.concat([result_df, pd.DataFrame([new_row])])


    if os.path.exists(result_file_path):
        # If the file already exists, load the existing data
        existing_df = pd.read_excel(result_file_path, sheet_name=sheet_name, engine='openpyxl')
        result_df = pd.concat([existing_df, result_df], ignore_index=True)

    with pd.ExcelWriter('data/regres/' + result_file, engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name=sheet_name, index=False)

        plot_sheet_name = func + '_' + phase
        plot_sheet = writer.book.create_sheet(title=plot_sheet_name)

        img_path = os.path.join(save_dir, 'plot.png')
        plot.savefig(img_path)

        img = Image(img_path)
        plot_sheet.add_image(img, 'A' + str(plot_sheet.max_row + 1))
        writer._save()


def regres_linear_analysis(input_file_path):
    df1 = pd.read_excel(input_file_path, sheet_name='meas_data', engine='openpyxl')
    df2 = pd.read_excel(input_file_path, sheet_name='calc_data', engine='openpyxl')
    merged_df = pd.merge(df1, df2, on='Имя файла')

    numeric_columns = merged_df.select_dtypes(include='number').columns
    merged_df = merged_df[numeric_columns].dropna(axis=1)

    columns_to_remove = [col for col in merged_df.columns if '_id' in col.lower()]
    merged_df = merged_df.drop(columns=columns_to_remove)

    # Добавление колонки q_total_lc
    merged_df['q_total_lc'] = merged_df['q_o_lc'] + merged_df['q_w_lc'] + merged_df['q_g_lc']

    # Формирование данных для обучения
    X = merged_df[['a_g', 'a_o', 'a_w', 'q_total_lc']]
    y_g = merged_df['q_g_lc']
    y_o = merged_df['q_o_lc']
    y_w = merged_df['q_w_lc']

    # Разделение на обучающий и тестовый наборы
    X_train, X_test, y_train_g, y_test_g, y_train_o, y_test_o, y_train_w, y_test_w = train_test_split(X, y_g, y_o, y_w, test_size=0.2, random_state=42)

    # Линейная регрессия для газа
    model_g = LinearRegression()
    model_g.fit(X_train, y_train_g)
    y_pred_g = model_g.predict(X_test)
    mse_g = mean_squared_error(y_test_g, y_pred_g)
    print(f'Mean Squared Error (газ): {mse_g}')
    print(f'Equation (газ): q_g_lc = {model_g.intercept_:.2f} + {model_g.coef_[0]:.2f}*a_g + {model_g.coef_[1]:.2f}*a_o + {model_g.coef_[2]:.2f}*a_w + {model_g.coef_[3]:.2f}*q_total_lc')

    # Линейная регрессия для exxsol
    model_o = LinearRegression()
    model_o.fit(X_train, y_train_o)
    y_pred_o = model_o.predict(X_test)
    mse_o = mean_squared_error(y_test_o, y_pred_o)
    print(f'Mean Squared Error (exxsol): {mse_o}')
    print(f'Equation (exxsol): q_o_lc = {model_o.intercept_:.2f} + {model_o.coef_[0]:.2f}*a_g + {model_o.coef_[1]:.2f}*a_o + {model_o.coef_[2]:.2f}*a_w + {model_o.coef_[3]:.2f}*q_total_lc')

    # Линейная регрессия для воды
    model_w = LinearRegression()
    model_w.fit(X_train, y_train_w)
    y_pred_w = model_w.predict(X_test)
    mse_w = mean_squared_error(y_test_w, y_pred_w)
    print(f'Mean Squared Error (вода): {mse_w}')
    print(f'Equation (вода): q_w_lc = {model_w.intercept_:.2f} + {model_w.coef_[0]:.2f}*a_g + {model_w.coef_[1]:.2f}*a_o + {model_w.coef_[2]:.2f}*a_w + {model_w.coef_[3]:.2f}*q_total_lc')

    # Проверка и создание директории для сохранения графиков
    save_dir = 'data/regres/linear'
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    save_dir_file = 'data/regres'
    # График для газа
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test_g, y_pred_g, color='blue')
    plt.plot([y_test_g.min(), y_test_g.max()], [y_test_g.min(), y_test_g.max()], linestyle='--', color='red',
             linewidth=2)
    plt.xlabel('Фактический расход газа')
    plt.ylabel('Предсказанный расход газа')
    plt.title('Линейная регрессия для газа')
    plt.savefig(os.path.join(save_dir, 'linear_regression_gas.png'))
    save_results_to_excel('gas', 'simple_linear', model_g, mse_g, save_dir_file, plt)
    plt.close()

    # График для exxsol
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test_o, y_pred_o, color='green')
    plt.plot([y_test_o.min(), y_test_o.max()], [y_test_o.min(), y_test_o.max()], linestyle='--', color='red',
             linewidth=2)
    plt.xlabel('Фактический расход exxsol')
    plt.ylabel('Предсказанный расход exxsol')
    plt.title('Линейная регрессия для exxsol')
    plt.savefig(os.path.join(save_dir, 'linear_regression_exxsol.png'))
    save_results_to_excel('exxsol', 'simple_linear',model_o, mse_o, save_dir_file, plt)
    plt.close()



    # График для воды
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test_w, y_pred_w, color='orange')
    plt.plot([y_test_w.min(), y_test_w.max()], [y_test_w.min(), y_test_w.max()], linestyle='--', color='red',
             linewidth=2)
    plt.xlabel('Фактический расход воды')
    plt.ylabel('Предсказанный расход воды')
    plt.title('Линейная регрессия для воды')
    plt.savefig(os.path.join(save_dir, 'linear_regression_water.png'))

    save_results_to_excel('water', 'simple_linear',model_w, mse_w, save_dir_file, plt)
    plt.close()

def equation_generate(func, coef):
    match func:
        case 'linear':
            return f'{coef[0]:.2f} * a_g + {coef[1]:.2f} * a_o + {coef[2]:.2f} * a_w + {coef[3]:.2f} * q_total_lc'
        case 'exp':
            return f'{coef[0]:.2f} * np.exp({coef[1]:.2f} * a_g + {coef[2]:.2f} * a_o + {coef[3]:.2f} * a_w + {coef[4]:.2f} * q_total_lc)'
        case 'quadratic':
            return f'{coef[0]:.2f} * a_g**2 + {coef[1]:.2f} * a_o**2 + {coef[2]:.2f} * a_w**2 + {coef[3]:.2f} * q_total_lc**2'
        case 'log':
            return f'{coef[0]:.2f} * np.log({coef[1]:.2f} * a_g + 1) + {coef[2]:.2f} * np.log({coef[3]:.2f} * a_o + 1) + {coef[4]:.2f} * np.log({coef[5]:.2f} * a_w + 1) + {coef[6]:.2f} * np.log({coef[7]:.2f} * q_total_lc + 1)'
        case 'simple_linear':
            return f'{coef.intercept_:.2f} + {coef.coef_[0]:.2f} * a_g + {coef.coef_[1]:.2f} * a_o + {coef.coef_[2]:.2f} * a_w + {coef.coef_[3]:.2f} * q_total_lc',
        case _:
            raise ValueError(f"Unsupported function type: {func}")


def non_linear_function(X, a, b, c, d, e, f, g, h, func):
    match func:
        case 'linear':
            return a * X[:, 0] + b * X[:, 1] + c * X[:, 2] + d * X[:, 3]
        case 'exp':
            return a * np.exp(b * X[:, 0] + c * X[:, 1] + d * X[:, 2] + e * X[:, 3])
        case 'quadratic':
            return a * X[:, 0]**2 + b * X[:, 1]**2 + c * X[:, 2]**2 + d * X[:, 3]**2
        case 'log':
            return a * np.log(b * X[:, 0] + 1) + c * np.log(d * X[:, 1] + 1) + e * np.log(f * X[:, 2] + 1) + g * np.log(h * X[:, 3] + 1)
        case _:
            raise ValueError(f"Unsupported function type: {func}")


def non_linear_regression_analysis(input_file_path, func):
    df1 = pd.read_excel(input_file_path, sheet_name='meas_data', engine='openpyxl')
    df2 = pd.read_excel(input_file_path, sheet_name='calc_data', engine='openpyxl')
    merged_df = pd.merge(df1, df2, on='Имя файла')

    numeric_columns = merged_df.select_dtypes(include='number').columns
    merged_df = merged_df[numeric_columns].dropna(axis=1)

    columns_to_remove = [col for col in merged_df.columns if '_id' in col.lower()]
    merged_df = merged_df.drop(columns=columns_to_remove)

    # Добавление колонки q_total_lc
    merged_df['q_total_lc'] = merged_df['q_o_lc'] + merged_df['q_w_lc'] + merged_df['q_g_lc']

    # Формирование данных для обучения
    X = merged_df[['a_g', 'a_o', 'a_w', 'q_total_lc']]
    y_g = merged_df['q_g_lc']
    y_o = merged_df['q_o_lc']
    y_w = merged_df['q_w_lc']

    # Разделение на обучающий и тестовый наборы
    X_train, X_test, y_train_g, y_test_g, y_train_o, y_test_o, y_train_w, y_test_w = train_test_split(X, y_g, y_o, y_w, test_size=0.2, random_state=42)

    # Нелинейная регрессия для газа
    popt_g, _ = curve_fit(lambda x, a, b, c, d, e, f, g, h: non_linear_function(x, a, b, c, d, e, f, g, h, func), X_train.values, y_train_g,
                          maxfev=10000)
    y_pred_g = non_linear_function(X_test.values, *popt_g, func)
    mse_g = mean_squared_error(y_test_g, y_pred_g)
    print(f'Mean Squared Error (газ): {mse_g}')
    print(f'Parameters (газ): {popt_g}')

    # Нелинейная регрессия для exxsol
    popt_o, _ = curve_fit(lambda x, a, b, c, d, e, f, g, h: non_linear_function(x, a, b, c, d, e, f, g, h, func), X_train.values, y_train_o, maxfev=10000)
    y_pred_o = non_linear_function(X_test.values, *popt_o, func)
    mse_o = mean_squared_error(y_test_o, y_pred_o)
    print(f'Mean Squared Error (exxsol): {mse_o}')
    print(f'Parameters (exxsol): {popt_o}')

    # Нелинейная регрессия для воды
    popt_w, _ = curve_fit(lambda x, a, b, c, d, e, f, g, h: non_linear_function(x, a, b, c, d, e, f, g, h, func), X_train.values, y_train_w, maxfev=10000)
    y_pred_w = non_linear_function(X_test.values, *popt_w, func)
    mse_w = mean_squared_error(y_test_w, y_pred_w)
    print(f'Mean Squared Error (вода): {mse_w}')
    print(f'Parameters (вода): {popt_w}')

    # Проверка и создание директории для сохранения графиков
    save_dir = 'data/regres/non_linear/' + func
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)


    save_dir_file = 'data/regres'


    # График для газа
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test_g, y_pred_g, color='blue')
    plt.plot([y_test_g.min(), y_test_g.max()], [y_test_g.min(), y_test_g.max()], linestyle='--', color='red',
             linewidth=2)
    plt.xlabel('Фактический расход газа')
    plt.ylabel('Предсказанный расход газа')
    plt.title('Нелинейная регрессия для газа')
    plt.savefig(os.path.join(save_dir, 'non_linear_' + func + '_gas.png'))
    save_results_to_excel('gas', func, popt_g, mse_g, save_dir_file, plt)
    plt.close()

    # График для exxsol
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test_o, y_pred_o, color='green')
    plt.plot([y_test_o.min(), y_test_o.max()], [y_test_o.min(), y_test_o.max()], linestyle='--', color='red',
             linewidth=2)
    plt.xlabel('Фактический расход exxsol')
    plt.ylabel('Предсказанный расход exxsol')
    plt.title('Нелинейная регрессия для exxsol')
    plt.savefig(os.path.join(save_dir, 'non_linear_' + func + '_exxsol.png'))
    save_results_to_excel('exxsol', func, popt_o, mse_o, save_dir_file, plt)
    plt.close()

    # График для воды
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test_w, y_pred_w, color='orange')
    plt.plot([y_test_w.min(), y_test_w.max()], [y_test_w.min(), y_test_w.max()], linestyle='--', color='red',
             linewidth=2)
    plt.xlabel('Фактический расход воды')
    plt.ylabel('Предсказанный расход воды')
    plt.title('Нелинейная регрессия для воды')
    plt.savefig(os.path.join(save_dir, 'non_linear_' + func + '_water.png'))
    save_results_to_excel('water', func, popt_w, mse_w, save_dir_file, plt)
    plt.close()

    # # Вывод уравнений
    # print(f'Equation (газ): q_g_lc = {popt_g[0]:.2f}*a_g + {popt_g[1]:.2f}*a_o + {popt_g[2]:.2f}*a_w + {popt_g[3]:.2f}*q_total_lc')
    # print(f'Equation (exxsol): q_o_lc = {popt_o[0]:.2f}*a_g + {popt_o[1]:.2f}*a_o + {popt_o[2]:.2f}*a_w + {popt_o[3]:.2f}*q_total_lc')
    # print(f'Equation (вода): q_w_lc = {popt_w[0]:.2f}*a_g + {popt_w[1]:.2f}*a_o + {popt_w[2]:.2f}*a_w + {popt_w[3]:.2f}*q_total_lc')


